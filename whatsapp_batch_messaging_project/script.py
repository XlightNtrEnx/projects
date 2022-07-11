from tokenize import String
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
import pandas
from openpyxl import load_workbook
import csv
 
excel_file_name = 'test.xlsx'

wb = load_workbook(filename = excel_file_name)
ws = wb.active

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
driver.get('https://web.whatsapp.com')

count = 0

message = "Testing"

input("Press ENTER after login into Whatsapp Web and your chats are visible.")

contact_list = []

for cell in tuple(ws.columns)[0]:
    if isinstance(cell.value, float): 
        contact_list.append(str(int(tuple(ws.columns)[0].value)))
    elif isinstance(cell.value, int):
        contact_list.append(str(tuple(ws.columns)[0].value))

print(len(contact_list))
print(contact_list)

for contact in contact_list:
    try:
        url = 'https://web.whatsapp.com/send?phone=+65' + contact + '&text=' + message
        sent = False
        driver.get(url)
        try:
            click_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'epia9gcq')))
        except Exception as e:
            print("Sorry message could not sent to " + contact)
            ws.cell(row=count+2, column=2).value = "No"
        else:
            sleep(4)
            click_btn.click()
            sent = True
            sleep(4)
            print('Message sent to: ' + contact)
            ws.cell(row=count+2, column=2).value = "Yes"
        finally:
            count = count + 1
    except Exception as e:
        print('Failed to send message to ' + contact + str(e))
        ws.cell(row=count+1, column=2).value = "No"
driver.quit()
wb.save(excel_file_name)
print("The script executed successfully.")